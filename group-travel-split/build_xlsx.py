from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
BLUE = "0000FF"   # input cứng
BLACK = "000000"  # công thức
GREEN = "008000"  # link sheet khác
HDR_FILL = PatternFill("solid", fgColor="1F3B4D")
BAND = PatternFill("solid", fgColor="EDF2F4")
NOTE_FILL = PatternFill("solid", fgColor="FFF9DB")
thin = Side(style="thin", color="B7C4CC")
medium = Side(style="medium", color="1F3B4D")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0;(#,##0);-'

def hdr(ws, row, labels, start=1):
    for i, t in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=t)
        c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX

def title(ws, text, sub=None):
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=ARIAL, size=14, bold=True, color="1F3B4D")
    if sub:
        s = ws.cell(row=2, column=1, value=sub)
        s.font = Font(name=ARIAL, size=9, italic=True, color="5A6B71")

wb = Workbook()

# ---------------------------------------------------------------- Chi phí
ws = wb["Sheet"]
ws.title = "Chi phí"
title(ws, "CHI PHÍ CHUYẾN ĐI", "Ô chữ xanh = số liệu nhập tay, ô chữ đen = công thức tự tính.")
hdr(ws, 4, ["Khoản chi", "Số tiền", "Chi trả", "Tú trả", "Việt trả", "Mỗi người"])

items = [
    ("Lưu trú",             13700000, 0,       0),
    ("Ăn trưa Cậu Khấu",     1228000, 0,       0),
    ("Ăn tối Hồng Phát",     2257400, 0,       0),
    ("Bar",                  1075200, 0,       0),
    ("Ăn sáng",               680000, 0,       0),
    ("Cafe sáng",             723600, 0,       0),
    ("Ăn tối BBQ",                 0, 1100000, 1000000),
    ("Ăn sáng 2",             645000, 0,       0),
    ("Snack",                 584445, 0,       0),
    ("Snack (thêm)",          180000, 0,       0),
]

r0 = 5
for i, (name, chi, tu, viet) in enumerate(items):
    r = r0 + i
    ws.cell(row=r, column=1, value=name).font = Font(name=ARIAL, size=10)
    ws.cell(row=r, column=2, value=f"=SUM(C{r}:E{r})").font = Font(name=ARIAL, size=10, bold=True)
    for col, val in ((3, chi), (4, tu), (5, viet)):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name=ARIAL, size=10, color=BLUE)
    ws.cell(row=r, column=6, value=f"=B{r}/'Chốt sổ'!$B$4").font = Font(name=ARIAL, size=10, color=GREEN)
    for col in range(1, 7):
        cell = ws.cell(row=r, column=col)
        cell.border = BOX
        if col > 1:
            cell.number_format = MONEY
        if i % 2 == 1:
            cell.fill = BAND

rt = r0 + len(items)
ws.cell(row=rt, column=1, value="TỔNG").font = Font(name=ARIAL, size=10, bold=True)
for col in range(2, 7):
    L = get_column_letter(col)
    c = ws.cell(row=rt, column=col, value=f"=SUM({L}{r0}:{L}{rt-1})")
    c.font = Font(name=ARIAL, size=10, bold=True)
    c.number_format = MONEY
for col in range(1, 7):
    cell = ws.cell(row=rt, column=col)
    cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
    cell.fill = PatternFill("solid", fgColor="DCE6EC")

n = ws.cell(row=rt + 2, column=1, value="Ghi chú: Toàn bộ các khoản do Chi thanh toán, trừ bữa BBQ do Tú trả 1.100.000 và Việt trả 1.000.000 (số liệu do người dùng cung cấp).")
n.font = Font(name=ARIAL, size=9, italic=True, color="5A6B71")
n.fill = NOTE_FILL

for col, w in zip("ABCDEF", (26, 15, 15, 14, 14, 15)):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------------------------------------------------------------- Chốt sổ
cs = wb.create_sheet("Chốt sổ")
title(cs, "CHỐT SỔ — CHIA ĐỀU 6 NGƯỜI", "Sửa số người / tiền cọc ở ô xanh, mọi kết quả bên dưới tự cập nhật.")

params = [
    ("Số người",                 6,                          BLUE),
    ("Tiền cọc mỗi người",       3000000,                    BLUE),
    ("Tổng quỹ cọc",             "=B4*B5",                   BLACK),
    ("Tổng chi phí",             "='Chi phí'!B15",           GREEN),
    ("Mỗi người phải chịu",      "=B7/B4",                   BLACK),
    ("Còn phải bù mỗi người",    "=B8-B5",                   BLACK),
]
for i, (label, val, color) in enumerate(params):
    r = 4 + i
    lc = cs.cell(row=r, column=1, value=label)
    lc.font = Font(name=ARIAL, size=10, bold=(i >= 4))
    vc = cs.cell(row=r, column=2, value=val)
    vc.font = Font(name=ARIAL, size=10, bold=(i >= 4), color=color)
    vc.number_format = "#,##0" if i else "0"
    for col in (1, 2):
        cs.cell(row=r, column=col).border = BOX
    if i >= 4:
        cs.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6EC")
        cs.cell(row=r, column=2).fill = PatternFill("solid", fgColor="DCE6EC")

HR = 12
hdr(cs, HR, ["Thành viên", "Cọc đã đóng", "Đã ứng ngoài quỹ", "Phải chịu",
             "Chênh lệch", "Chốt (làm tròn 100đ)", "Kết luận"])

# Chi ở dòng đầu để công thức plug -SUM(F14:F18) gom phần lẻ làm tròn về Chi
members = [
    ("Chi",   "='Chi phí'!C15-$B$6"),
    ("Quỳnh", 0),
    ("Hà",    0),
    ("Hiếu",  0),
    ("Việt",  "='Chi phí'!E15"),
    ("Tú",    "='Chi phí'!D15"),
]
m0 = HR + 1
for i, (name, advance) in enumerate(members):
    r = m0 + i
    cs.cell(row=r, column=1, value=name).font = Font(name=ARIAL, size=10, bold=True)
    cs.cell(row=r, column=2, value="=$B$5").font = Font(name=ARIAL, size=10)
    ac = cs.cell(row=r, column=3, value=advance)
    ac.font = Font(name=ARIAL, size=10, color=GREEN if isinstance(advance, str) else BLUE)
    cs.cell(row=r, column=4, value="=$B$8").font = Font(name=ARIAL, size=10)
    cs.cell(row=r, column=5, value=f"=B{r}+C{r}-D{r}").font = Font(name=ARIAL, size=10)
    if i == 0:
        cs.cell(row=r, column=6, value=f"=-SUM(F{m0+1}:F{m0+len(members)-1})")
    else:
        cs.cell(row=r, column=6, value=f"=ROUND(E{r},-2)")
    cs.cell(row=r, column=6).font = Font(name=ARIAL, size=10, bold=True)
    cs.cell(row=r, column=7, value=f'=IF(F{r}>0,"Nhận về","Trả thêm")').font = Font(name=ARIAL, size=10)
    for col in range(1, 8):
        cell = cs.cell(row=r, column=col)
        cell.border = BOX
        if 2 <= col <= 6:
            cell.number_format = MONEY
        if col == 7:
            cell.alignment = Alignment(horizontal="center")
        if i % 2 == 1:
            cell.fill = BAND

rt2 = m0 + len(members)
cs.cell(row=rt2, column=1, value="TỔNG").font = Font(name=ARIAL, size=10, bold=True)
for col in range(2, 7):
    L = get_column_letter(col)
    c = cs.cell(row=rt2, column=col, value=f"=SUM({L}{m0}:{L}{rt2-1})")
    c.font = Font(name=ARIAL, size=10, bold=True)
    c.number_format = MONEY
cs.cell(row=rt2, column=7, value='=IF(ROUND(E' + str(rt2) + ',2)=0,"Cân sổ","LỆCH!")').font = Font(name=ARIAL, size=10, bold=True)
cs.cell(row=rt2, column=7).alignment = Alignment(horizontal="center")
for col in range(1, 8):
    cell = cs.cell(row=rt2, column=col)
    cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
    cell.fill = PatternFill("solid", fgColor="DCE6EC")

notes = [
    "CÁCH TÍNH",
    "• Giả định: tiền cọc 3.000.000/người được gom thành quỹ chung do Chi giữ và Chi dùng quỹ này để thanh toán.",
    "• 'Đã ứng ngoài quỹ' của Chi = tổng Chi thanh toán (21.073.645) − tổng quỹ cọc (18.000.000) = 3.073.645 tiền túi.",
    "• Tú và Việt trả bữa BBQ trực tiếp nên toàn bộ số đó là tiền ứng ngoài quỹ.",
    "• Chênh lệch = Cọc đã đóng + Đã ứng ngoài quỹ − Phải chịu. Dương = được nhận lại, âm = phải trả thêm.",
    "• Cột 'Chốt' làm tròn đến 100đ; phần lẻ do làm tròn được gom vào dòng của Chi để tổng thu và tổng chi luôn khớp.",
    "• Nếu tiền cọc thực tế nộp cho bên khác (không phải Chi giữ), hãy sửa lại cột 'Đã ứng ngoài quỹ' cho đúng.",
]
for i, t in enumerate(notes):
    c = cs.cell(row=rt2 + 2 + i, column=1, value=t)
    c.font = Font(name=ARIAL, size=9, bold=(i == 0), italic=(i > 0), color="1F3B4D" if i == 0 else "5A6B71")
    if i > 0:
        c.fill = NOTE_FILL

for col, w in zip("ABCDEFG", (24, 18, 19, 15, 15, 20, 13)):
    cs.column_dimensions[col].width = w
cs.freeze_panes = "A13"

# ---------------------------------------------------------------- Chuyển tiền
ct = wb.create_sheet("Chuyển tiền")
title(ct, "CHUYỂN TIỀN — 5 LƯỢT LÀ XONG", "Số tiền lấy trực tiếp từ cột 'Chốt' của sheet Chốt sổ.")
hdr(ct, 4, ["#", "Người chuyển", "Người nhận", "Số tiền"])

# Chốt sổ rows: 13 Chi, 14 Quỳnh, 15 Hà, 16 Hiếu, 17 Việt, 18 Tú
transfers = [
    ("Quỳnh", "Chi",  "=-'Chốt sổ'!F14"),
    ("Hà",    "Chi",  "=-'Chốt sổ'!F15"),
    ("Hiếu",  "Tú",   "='Chốt sổ'!F18"),
    ("Hiếu",  "Việt", "='Chốt sổ'!F17"),
    ("Hiếu",  "Chi",  "=-'Chốt sổ'!F16-D7-D8"),
]
for i, (a, b, f) in enumerate(transfers):
    r = 5 + i
    ct.cell(row=r, column=1, value=i + 1).font = Font(name=ARIAL, size=10, color="5A6B71")
    ct.cell(row=r, column=2, value=a).font = Font(name=ARIAL, size=10, bold=True)
    ct.cell(row=r, column=3, value=b).font = Font(name=ARIAL, size=10)
    c = ct.cell(row=r, column=4, value=f)
    c.font = Font(name=ARIAL, size=10, bold=True, color=GREEN)
    c.number_format = MONEY
    for col in range(1, 5):
        cell = ct.cell(row=r, column=col)
        cell.border = BOX
        if i % 2 == 1:
            cell.fill = BAND
    ct.cell(row=r, column=1).alignment = Alignment(horizontal="center")

rt3 = 5 + len(transfers)
ct.cell(row=rt3, column=2, value="TỔNG").font = Font(name=ARIAL, size=10, bold=True)
c = ct.cell(row=rt3, column=4, value=f"=SUM(D5:D{rt3-1})")
c.font = Font(name=ARIAL, size=10, bold=True)
c.number_format = MONEY
for col in range(1, 5):
    cell = ct.cell(row=rt3, column=col)
    cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
    cell.fill = PatternFill("solid", fgColor="DCE6EC")

chk = [
    ("Chi nhận đủ",  "=D5+D6+D9"),
    ("Tú nhận đủ",   "=D7"),
    ("Việt nhận đủ", "=D8"),
    ("Hiếu chuyển đi tổng", "=D7+D8+D9"),
]
ct.cell(row=rt3 + 2, column=2, value="KIỂM TRA").font = Font(name=ARIAL, size=9, bold=True, color="1F3B4D")
for i, (label, f) in enumerate(chk):
    r = rt3 + 3 + i
    ct.cell(row=r, column=2, value=label).font = Font(name=ARIAL, size=9, italic=True, color="5A6B71")
    c = ct.cell(row=r, column=4, value=f)
    c.font = Font(name=ARIAL, size=9, color="5A6B71")
    c.number_format = MONEY

for col, w in zip("ABCD", (6, 18, 18, 18)):
    ct.column_dimensions[col].width = w

wb.save("/home/user/brainstorm/group-travel-split/chia-tien-chuyen-di.xlsx")
print("saved")
